from openpyxl import (
    load_workbook, 
    Workbook
)
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    NamedStyle,
    Protection
)
from openpyxl.worksheet.page import (
    PageMargins, 
    PrintOptions,
    PrintPageSetup,
    
    )
from openpyxl.worksheet.views import (
    SheetViewList,
    SheetView
)
from copy import copy
from openpyxl.utils.cell import column_index_from_string
import re
import pandas as pd

def amount_converter(data:float):
    '''
    小写金额转大写金额
    data: float scalar
    '''
    def _csplit(cdata):  # 拆分函数，将整数字符串拆分成[亿，万，仟]的list
        '''
        将整数部分字符数字以4个为一组进行分割，返回 list
        例：csplit('2938329803') 返回 ['29', '3832', '9803']
        '''                      
        g = len(cdata) % 4
        csdata = []
        lx = len(cdata) - 1
        if g > 0:
            csdata.append(cdata[0:g])
        k = g
        while k <= lx:
            csdata.append(cdata[k:k + 4])
            k += 4
        return csdata

    def _cschange(cki):  # 对[亿，万，仟]的list中每个字符串分组进行大写化再合并
        '''
        将4位字符转为仟佰十的大写
        例： cschange('9803'), 返回 '玖仟捌佰零叁'
        '''
        lenki = len(cki)
        i = 0
        lk = lenki
        chk = u''
        for i in range(lenki):
            if int(cki[i]) == 0:
                if i < lenki - 1:
                    if int(cki[i + 1]) != 0:
                        chk = chk + _gdict[int(cki[i])]
            else:
                chk = chk + _gdict[int(cki[i])] + _cdict[lk]
            lk -= 1
        return chk

    _cdict = {1: u'', 2: u'拾', 3: u'佰', 4: u'仟'}
    _xdict = {1: u'元', 2: u'万', 3: u'亿', 4: u'兆'}  # 数字标识符
    _gdict = {0: u'零', 1: u'壹', 2: u'贰', 3: u'叁', 4: u'肆', 5: u'伍', 6: u'陆', 7: u'柒', 8: u'捌', 9: u'玖'}

    data = float(data)
    cdata = str(data).split('.')

    cki = cdata[0]  # 整数部分
    ckj = cdata[1]  # 小数部分
    i = 0
    chk = u''
    cski = _csplit(cki)  # 分解字符数组[亿，万，仟]三组List:['0000','0000','0000']
    ikl = len(cski)  # 获取拆分后的List长度
    # 大写合并
    for i in range(ikl):
        if _cschange(cski[i]) == '':  # 有可能一个字符串全是0的情况
            chk = chk + _cschange(cski[i])  # 此时不需要将数字标识符引入
        else:
            chk = chk + _cschange(cski[i]) + _xdict[ikl - i]  # 合并：前字符串大写+当前字符串大写+标识符
    # 处理小数部分
    lenkj = len(ckj)
    if lenkj == 1:  # 若小数只有1位
        if int(ckj[0]) == 0:
            chk = chk + u''
        else:
            chk = chk + _gdict[int(ckj[0])] + u'角'
    else:  # 若小数有两位的四种情况
        if int(ckj[0]) == 0 and int(ckj[1]) != 0:
            chk = chk + u'零' + _gdict[int(ckj[1])] + u'分'
        elif int(ckj[0]) == 0 and int(ckj[1]) == 0:
            chk = chk + u''
        elif int(ckj[0]) != 0 and int(ckj[1]) != 0:
            chk = chk + _gdict[int(ckj[0])] + u'角' + _gdict[int(ckj[1])] + u'分'
        else:
            chk = chk + _gdict[int(ckj[0])] + u'角'
    return chk if chk[-1:]=='分' else chk+'整'

class qhcell:
    value = None  # 内容
    style = None  
    row = None    # 所在行
    col = None    # 所在列
    def __init__(self, value=None, style=None,row=None,col=None):
        self.value = value
        self.style = style
        self.row = row
        self.col = col
        
class qhblock:
    heights = []  # 所有 cell's height 的集合体
    widths = []   # 所有 cells' width 的集合体
    matrix = None # 所有 qhcell 的集合体
    merges = []   # 所有合并块的集合体， 如 ['A1:B2', "C4:C6"]
    
    def __init__(self, heights=None, widths=None, matrix=None, merges=None):
        self.heights = heights
        self.widths = widths
        self.matrix = matrix
        self.merges = merges
        self.get_block_height()

    def get_block_height(self):
        the_number_of_lines_in_the_block = None
        if(self.matrix):
            the_number_of_lines_in_the_block = len(self.matrix)            
        return the_number_of_lines_in_the_block
    
    def get_block_width(self):
        the_number_of_columns_in_the_block = None
        if(self.matrix):
            the_number_of_columns_in_the_block = len(self.matrix[0])            
        return the_number_of_columns_in_the_block   
    
class pattern_copier:
    NARROW = {
        'top':0.75, 'bottom':0.75,
        'left':0.25, 'right':0.25,
        'header':0.3, 'footer':0.3
    }
    WIDE = {
        'top':1, 'bottom':1,
        'left':1, 'right':1,
        'header':0.5, 'footer':0.5     
    }
    NORMAL = {
        'top':0.75, 'bottom':0.75,
        'left':0.7, 'right':0.7,
        'header':0.3, 'footer':0.3
    }

    def __init__(self, pattern_file, data_file, out_file, filling_loc,  pattern_file_idx=0, data_file_idx=0):
        self.pattern_file = pattern_file
        self.pattern_file_idx = pattern_file_idx
        self.data_file = data_file
        self.data_file_idx = data_file_idx
        self.out_file = out_file
        self.filling_loc = filling_loc
        self.filling_datas = pd.read_excel(self.data_file)
        self.block = self._getblock()
        self.outbook = Workbook()
        self.outsheet = self.outbook.active

    def _getblock(self):
        '''
        从指定excel文件的指定sheet中，获取模板数据块
        file: excel文件名
        sheet: worksheet名
        '''
        wb = load_workbook(self.pattern_file)
        ws = wb.worksheets[self.pattern_file_idx]
        
        matrix = []
        for row in range(1, ws.max_row+1):
            cellsinrow = []
            for col in range(1, ws.max_column+1):        
                value = ws.cell(row,col).value
                style = NamedStyle(name=f'r{row}c{col}')
                style.font = copy(ws.cell(row,col).font)
                style.alignment = copy(ws.cell(row,col).alignment)
                style.border = copy(ws.cell(row,col).border)
                style.number_format = copy(ws.cell(row,col).number_format)
                cell = qhcell(value,style,row,col)
                cellsinrow.append(cell)
            matrix.append(cellsinrow)

        heights = []
        for row in range(1, ws.max_row+1):
            heights.append(
                ws.row_dimensions[row].height
            )

        widths = []
        for col in range(1, ws.max_column+1):

            widths.append(
                ws.column_dimensions[
                    get_column_letter(col)
                ].width
            )

        merges = [rng.coord for rng in ws.merged_cells.ranges]
        
        wb.close()

        block = qhblock(heights, widths, matrix, merges)   
        return block
    
    def _copyblock(self, filling_data, offset_row=0, offset_col=0):
        '''
        将模板 block 复制一份到 worksheet 表中，复制的起始位置从（1，1）偏移（offset_row, offset_col),
        并按给定的位置 filling_loc 填充数据 data。
        eg. 
            loc = {
                'M2':'填单日期',
                'E3':'收款单位', 'M3':'请款部门',
                'E4':'银行账号', 'M4':'开户银行',
                'F5':'金额（大写）','M5':'转款金额', 
                'B7':'用途', 
                'O9':'单据数量'
            }        
            rec = pd.Series(
                data=['2021-06-26', 
                 '王六', 
                 None, 
                 '888 888 888 888',
                 '中信银行', 234614.83, 
                 '收购废钢款', 
                 2, 
                 '贰拾叁万肆仟陆佰壹拾肆元捌角叁分'
                ],
                index=['填单日期', '收款单位', '请款部门', '银行账号', '开户银行', '转款金额', '用途', '单据数量', '金额（大写）']
            )           
            copyblock(ws, block, 0,5,loc,rec)
        '''
        def _convert_cell(cellstr, offset_row, offset_col):
            '''
            将单元格字符地址转换为偏移后新的字符地址
            eg. call _convert_cell('B5', 2, 2)   return 'D7'
            '''
            col, row = re.match(r'([a-zA-Z]+)([0-9]+)', cellstr).groups()
            col = get_column_letter(
                column_index_from_string(col) + offset_col
            )
            row = int(row) + offset_row   
            return f'{col}{row}'

        def convert_merged_range(merged_range, offset_row, offset_col):
            '''
            将合并范围字符转换为偏移后的新范围字符
            example: call convert_merged_range('A1:B2',3,3), return 'D4:E5'
            '''
            loc = []
            rng = merged_range.split(':')
            for i in range(len(rng)):                
                loc.append(
                    _convert_cell(rng[i], offset_row, offset_col)
                )
            return ':'.join(loc)       
               
        ws = self.outsheet
        
        # setting the height of all rows in the block 
        for row in range(offset_row+1, offset_row+len(self.block.heights)+1):           
            ws.row_dimensions[row].height = self.block.heights[row-offset_row-1]

        # setting the width of all columns in the block
        for col in range(offset_col+1, offset_col+len(self.block.widths)+1):
            ws.column_dimensions[
                get_column_letter(col)
            ].width = self.block.widths[col-offset_col-1]

        # filling the value of all cells in block
        for row in self.block.matrix:
            for cell in row:
                ws.cell(offset_row+cell.row, offset_col+cell.col).value = cell.value
                ws.cell(offset_row+cell.row, offset_col+cell.col).style = cell.style
        
        # merging each range in the block
        for item in self.block.merges:
            ws.merge_cells(
                convert_merged_range(item, offset_row, offset_col)
            )
            
        # fill in the data automatically
        for cell_loc in self.filling_loc:           
            offseted_loc = _convert_cell(cell_loc, offset_row, offset_col)            
            ws[offseted_loc].value = filling_data[self.filling_loc[cell_loc]]                    
            
        return ws

    def addcnamount(self,cn_colname, al_colname):
        self.filling_datas[cn_colname] = self.filling_datas[al_colname].apply(amount_converter) 

    def filterdatas(self, colname, colvalues:list, *args):
        self.filling_datas = self.filling_datas.reset_index(drop=True)
        if colname!='' and colvalues is not None:
            self.filling_datas = self.filling_datas[
                self.filling_datas[colname].isin(colvalues)
            ]
            self.filling_datas.reset_index(drop=True, inplace=True)

    def copyblocks(self, start_row=1, start_col=1,
        number_of_blocks_per_page=1,
        gap_between_blocks_in_page=1,
        gap_between_pages=2,
        ):
        '''
        按 datas 提供的记录条数，复制对应数量的 block, 并填上相应信息
        '''
        first_offset_row = start_row - 1
        first_offset_col = start_col - 1
        block_height = self.block.get_block_height()

        page_height = (
            block_height * number_of_blocks_per_page + gap_between_blocks_in_page * (number_of_blocks_per_page - 1)
            + gap_between_pages
        )       
        
        for i, row in self.filling_datas.iterrows():
            # i -= 1  # based-zero
            no_block_in_page = i % number_of_blocks_per_page
            no_page = int((i - no_block_in_page) / number_of_blocks_per_page)
            offset_row = first_offset_row + no_page * page_height + no_block_in_page * (block_height + gap_between_blocks_in_page)
            offset_col = first_offset_col
            
            self._copyblock(
                filling_data=row,
                offset_row=offset_row, 
                offset_col=offset_col               
            )

            # print(no_block_in_page, no_page, offset_row)

            # to add a row break at the bottom of each page 
            if ((no_block_in_page + 1) == number_of_blocks_per_page):        
                last_row_in_page = offset_row + block_height + gap_between_pages
                self.outsheet.row_breaks.append(Break(last_row_in_page))

        # to set print area
        print_area = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(self.outsheet.max_column)}{self.outsheet.max_row}'
        self.outsheet.print_area = print_area

    def pagelayout(self, margins=NORMAL, scale=100, paperSize='0', orientation='portrait', horizontalcentered=True, verticalcentered=True ):

        self.outsheet.page_margins = PageMargins(
            left=margins['left'], 
            right=margins['right'], 
            top=margins['top'], 
            bottom=margins['bottom'],
            header=margins['header'],
            footer=margins['footer']
        )
        self.outsheet.page_setup = PrintPageSetup(
            self.outsheet,
            scale=scale,
            paperSize=paperSize,
            orientation=orientation,
        )            
        self.outsheet.print_options = PrintOptions(
            horizontalCentered=horizontalcentered,
            verticalCentered=verticalcentered
        )

        self.outsheet.views = SheetViewList([SheetView(showGridLines=False, showRowColHeaders=False)])

    def output(self):
        self.outbook.save(self.out_file)
        self.outbook.close()
        print(f'haved outputed in {self.out_file} file.')